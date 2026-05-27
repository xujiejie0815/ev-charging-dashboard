# -*- coding: utf-8 -*-
"""
EV充電 利用状況 RAW DATA CSV生成スクリプト
- 全施設×全月（稼働開始日〜稼働終了日 or 当月）
- 使用なし月は稼働時間=0でゼロ埋め
- GitHub Actions: XLSX_PATH / OUT_PATH を環境変数で上書き可
"""
import os
import sys
import urllib.request
import openpyxl
import pandas as pd
from datetime import datetime, date
import calendar
import re

SHEETS_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "19f1cwAWaIncAwns4JA7uw_ZGzKXVydWYZw66jMuapRY"
    "/export?format=xlsx"
)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH  = os.environ.get("XLSX_PATH", "/tmp/spreadsheet_ev.xlsx")
OUT_PATH   = os.environ.get(
    "OUT_PATH",
    os.path.join(SCRIPT_DIR, "data", "all_facilities_usage_v5.csv"),
)

_now = datetime.now()
END_YM = (_now.year, _now.month)

# ── 0. XLSX ダウンロード（ファイルが存在しない場合） ─────────────────────
if not os.path.exists(XLSX_PATH):
    print(f"Downloading XLSX from Google Sheets → {XLSX_PATH}")
    urllib.request.urlretrieve(SHEETS_URL, XLSX_PATH)
    print("  Download complete.")
else:
    print(f"Using existing XLSX: {XLSX_PATH}")

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

# ── 1. 施設P マスタ読み込み ──────────────────────────────────────
ws_fac = wb['施設P']
fac_rows = list(ws_fac.iter_rows(values_only=True))

facilities = {}  # gid_str -> dict
for r in fac_rows[1:]:
    if r[0] is None:
        continue
    gid = str(int(r[0]))
    release_dt = r[3]
    if isinstance(release_dt, datetime):
        release_ym = (release_dt.year, release_dt.month)
    else:
        release_ym = None
    facilities[gid] = {
        'gid':        gid,
        'fac_id':     str(int(r[1])) if r[1] is not None else gid,
        'model':      r[2] or '',
        'release_dt': release_dt,
        'release_ym': release_ym,
        'operator':   r[4] or '',
        'category':   r[5] or '',
        'brand':      r[6] or '',
        'name':       r[7] or '',
        'units':      int(r[8]) if r[8] is not None else 1,
    }

print(f"施設P: {len(facilities)} facilities loaded")

# ── 2. 利用履歴 読み込み（primary） ────────────────────────────────
ws_hist = wb['利用履歴']
hist_rows = list(ws_hist.iter_rows(values_only=True))
hist_headers = hist_rows[0]

def parse_ym_str(s):
    if s is None:
        return None
    m = re.match(r'(\d{4})-(\d+)月', str(s))
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return None

def parse_ym_dt(dt):
    if isinstance(dt, datetime):
        return (dt.year, dt.month)
    return None

def parse_end_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return (v.year, v.month)
    if isinstance(v, str) and v.strip():
        try:
            dt = datetime.strptime(v.strip()[:10], '%Y/%m/%d')
            return (dt.year, dt.month)
        except Exception:
            pass
    return None

usage = {}  # gid_str → dict{(y,m): row}
hist_col = {h: i for i, h in enumerate(hist_headers) if h is not None}

for r in hist_rows[1:]:
    gid_raw = r[hist_col.get('充電器グループID', 0)]
    if not gid_raw:
        continue
    gid = str(gid_raw)
    ym = parse_ym_str(r[hist_col['利用月']])
    if ym is None:
        continue
    if gid not in usage:
        usage[gid] = {}
    usage[gid][ym] = {
        '日数':              r[hist_col['日数']],
        '台数':              r[hist_col['台数']],
        '稼働開始日':        r[hist_col['稼働開始日']],
        '稼働終了日(手動)':  r[hist_col['稼働終了日(手動)']],
        '稼働時間(分)':      r[hist_col['稼働時間(分)']],
        '利用人数':          r[hist_col['利用人数']],
        '利用回数':          r[hist_col['利用回数']],
        '稼働率':            r[hist_col['稼働率']],
        '1人あたりの利用分数': r[hist_col['1人あたりの利用分数']],
        '1回あたりの利用分数': r[hist_col['1回あたりの利用分数']],
        '平均利用頻度':      r[hist_col['平均利用頻度']],
    }

print(f"利用履歴: {len(usage)} GIDs, "
      f"{sum(len(v) for v in usage.values())} rows loaded")

# ── 3. RAWDATA 読み込み（secondary / fallback） ──────────────────────
raw_usage = {}

if '【RAWDATA】all_facilities_usage' in wb.sheetnames:
    ws_raw = wb['【RAWDATA】all_facilities_usage']
    raw_rows = list(ws_raw.iter_rows(values_only=True))
    raw_headers = raw_rows[0]
    raw_col = {h: i for i, h in enumerate(raw_headers) if h is not None}

    for r in raw_rows[1:]:
        gid_raw = r[raw_col.get('充電器グループID', 0)]
        if gid_raw is None:
            continue
        gid = str(int(gid_raw))
        ym_raw = r[raw_col['利用月']]
        ym = parse_ym_dt(ym_raw) if isinstance(ym_raw, datetime) else parse_ym_str(ym_raw)
        if ym is None:
            continue
        if gid not in raw_usage:
            raw_usage[gid] = {}
        raw_usage[gid][ym] = {
            '日数':              r[raw_col['日数']],
            '台数':              r[raw_col['台数']],
            '稼働開始日':        r[raw_col['稼働開始日']],
            '稼働終了日(手動)':  r[raw_col['稼働終了日(手動)']],
            '稼働時間(分)':      r[raw_col['稼働時間(分)']],
            '利用人数':          r[raw_col['利用人数']],
            '利用回数':          r[raw_col['利用回数']],
            '稼働率':            r[raw_col['稼働率']],
            '1人あたりの利用分数': r[raw_col['1人あたりの利用分数']],
            '1回あたりの利用分数': r[raw_col['1回あたりの利用分数']],
            '平均利用頻度':      r[raw_col['平均利用頻度']],
        }
    print(f"RAWDATA: {len(raw_usage)} GIDs, "
          f"{sum(len(v) for v in raw_usage.values())} rows loaded")
else:
    print("RAWDATA: シートなし（スキップ）")

# ── 4. 各施設の稼働終了日を収集 ──────────────────────────────────────
def get_end_ym_for_gid(gid, src_usage):
    end_ym_found = None
    for ym, row in src_usage.get(gid, {}).items():
        e = parse_end_date(row.get('稼働終了日(手動)'))
        if e and (end_ym_found is None or e < end_ym_found):
            end_ym_found = e
    return end_ym_found

def get_end_dt_for_gid(gid, src_usage):
    end_dt_found = None
    for ym, row in src_usage.get(gid, {}).items():
        v = row.get('稼働終了日(手動)')
        if v is None:
            continue
        if isinstance(v, datetime):
            dt = v
        elif isinstance(v, str) and v.strip():
            try:
                dt = datetime.strptime(v.strip()[:10], '%Y/%m/%d')
            except Exception:
                continue
        else:
            continue
        if end_dt_found is None or dt < end_dt_found:
            end_dt_found = dt
    return end_dt_found

# ── 5. 全施設×全月Rows生成 ──────────────────────────────────────
def ym_to_str(y, m):
    return f"{y}-{m}月"

def days_in_month(y, m):
    return calendar.monthrange(y, m)[1]

def dt_to_str(dt):
    if isinstance(dt, datetime):
        return dt.strftime('%Y/%m/%d')
    return '' if dt is None else str(dt)

def next_ym(y, m):
    m += 1
    if m > 12:
        m = 1
        y += 1
    return y, m

OUTPUT_COLS = [
    '充電器グループID', '施設ID', '施設名', 'モデル', 'カテゴリー',
    '利用月', '日数', '台数', '稼働開始日', '稼働終了日(手動)',
    '稼働時間(分)', '利用人数', '利用回数', '稼働率',
    '1人あたりの利用分数', '1回あたりの利用分数', '平均利用頻度',
    'ブランド', '運営会社',
]

all_rows = []

for gid, fac in sorted(facilities.items(), key=lambda x: int(x[0])):
    if gid in usage:
        src = usage[gid]
    elif gid in raw_usage:
        src = raw_usage[gid]
    else:
        src = {}

    release_ym = fac['release_ym']
    if release_ym is None:
        if src:
            release_ym = min(src.keys())
        else:
            continue

    end_ym_for_gid = get_end_ym_for_gid(gid, usage)
    if end_ym_for_gid is None:
        end_ym_for_gid = get_end_ym_for_gid(gid, raw_usage)
    cap_end = end_ym_for_gid if end_ym_for_gid else END_YM
    cap_end = min(cap_end, END_YM)

    end_dt_full = get_end_dt_for_gid(gid, usage)
    if end_dt_full is None:
        end_dt_full = get_end_dt_for_gid(gid, raw_usage)

    end_dt_str = ''
    for ym_key, row in src.items():
        v = parse_end_date(row.get('稼働終了日(手動)'))
        if v:
            e_val = row['稼働終了日(手動)']
            end_dt_str = dt_to_str(e_val)
            break

    if src:
        earliest_actual = min(src.keys())
    else:
        earliest_actual = release_ym

    start_ym_gen = min(earliest_actual, release_ym)
    rel_str = dt_to_str(fac['release_dt'])

    y, m = start_ym_gen
    while (y, m) <= cap_end:
        ym = (y, m)
        ym_str = ym_to_str(y, m)
        d = days_in_month(y, m)

        if ym in src:
            row = src[ym]
            ops_min   = float(row['稼働時間(分)'])   if row['稼働時間(分)']   is not None else 0.0
            pax       = float(row['利用人数'])         if row['利用人数']         is not None else 0.0
            trips     = float(row['利用回数'])         if row['利用回数']         is not None else 0.0
            units_row = float(row['台数'])             if row['台数']             is not None else float(fac['units'])
            days_row  = float(row['日数'])             if row['日数']             is not None else float(d)
            rate_raw  = row['稼働率']
            min_per_p = row['1人あたりの利用分数']
            min_per_t = row['1回あたりの利用分数']
            avg_freq  = row['平均利用頻度']
        else:
            ops_min   = 0.0
            pax       = 0.0
            trips     = 0.0
            units_row = float(fac['units'])
            days_row  = float(d)
            rate_raw  = None
            min_per_p = None
            min_per_t = None
            avg_freq  = None

        if end_dt_full and (y, m) == (end_dt_full.year, end_dt_full.month):
            effective_days = max(0.0, float(end_dt_full.day - 1))
        else:
            effective_days = days_row

        possible = effective_days * 24 * 60 * units_row
        if possible > 0 and ops_min > 0:
            rate_calc = round(max(0.0, ops_min / possible * 100), 4)
        else:
            rate_calc = 0.0

        all_rows.append({
            '充電器グループID':     gid,
            '施設ID':               fac['fac_id'],
            '施設名':               fac['name'],
            'モデル':               fac['model'],
            'カテゴリー':           fac['category'],
            '利用月':               ym_str,
            '日数':                 effective_days,
            '台数':                 units_row,
            '稼働開始日':           rel_str,
            '稼働終了日(手動)':     end_dt_str,
            '稼働時間(分)':         round(ops_min, 4),
            '利用人数':             int(pax),
            '利用回数':             int(trips),
            '稼働率':               rate_calc,
            '1人あたりの利用分数':  round(float(min_per_p), 4) if min_per_p is not None else 0.0,
            '1回あたりの利用分数':  round(float(min_per_t), 4) if min_per_t is not None else 0.0,
            '平均利用頻度':         round(float(avg_freq), 4)  if avg_freq  is not None else 0.0,
            'ブランド':             fac['brand'],
            '運営会社':             fac['operator'],
        })

        y, m = next_ym(y, m)

# ── 6. DataFrame → CSV ────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
df = pd.DataFrame(all_rows, columns=OUTPUT_COLS)
df.to_csv(OUT_PATH, index=False, encoding='utf-8-sig')

print(f"\n✅ CSV saved: {OUT_PATH}")
print(f"   Facilities: {df['充電器グループID'].nunique()}")
print(f"   Total rows: {len(df)}")
print(f"   Zero rows:  {(df['稼働時間(分)'] == 0).sum()}")
print(f"   Non-zero rows: {(df['稼働時間(分)'] > 0).sum()}")
print(f"   稼働時間(分) total: {df['稼働時間(分)'].sum():,.2f}")
print(f"   Terminated GIDs: {(df['稼働終了日(手動)'] != '').sum()} rows")
